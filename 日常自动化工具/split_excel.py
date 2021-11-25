import xlrd
from openpyxl import Workbook, load_workbook

file_name = '1.xlsx'

data = xlrd.open_workbook(file_name)
table_0 = data.sheet_by_index(0)
result_path = file_name
wb = load_workbook(result_path)
ws = wb['Sheet1']


rows_num  = table_0.nrows - 1

data_list = []

for temp_rows in range(rows_num):
    temp_data =  table_0.cell(temp_rows+1,0).value
    temp_list = temp_data.split(" ")
    true_temp_list = []
    for jj in range(len(temp_list)):
        temp_list[jj] = temp_list[jj].strip()
        if temp_list[jj]!= "":
            #不能为空
            true_temp_list.append(temp_list[jj])
    data_list.append(true_temp_list)

#写入
for temp_rows in range(rows_num):
    for temp_lines in range(len(data_list[temp_rows])):
        # ws.cell(row=temp_rows + 2, column=temp_lines + 1).value = data_list[temp_rows][temp_lines]
        # 大写
        ws.cell(row=temp_rows + 2, column=temp_lines + 1).value = data_list[temp_rows][temp_lines].upper()

wb.save(file_name)
print("ok")


